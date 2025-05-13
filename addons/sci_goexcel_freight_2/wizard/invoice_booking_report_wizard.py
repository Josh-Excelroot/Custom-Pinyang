from odoo import api, fields, models
from datetime import date,datetime
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
import tempfile
import base64
import copy
import locale


class InvoiceBookingReportWizard(models.TransientModel):
    _name = "invoice.booking.report.wizard"

    date_from = fields.Date(
        string='Date From',
        required=True,
    )
    date_to = fields.Date(
        string='Date To',
        required=True,
    )

    file_data = fields.Binary(string="File Data", readonly=True)
    file_name = fields.Char(string="File Name", readonly=True)

    # Get month list
    def get_month_unique(self, invoice_list):
        unique_month = []
        for inv in invoice_list:
            month_year = {
                'month': inv.date_invoice.strftime('%b'),  # 'Jan', 'Feb', etc.
                'year': inv.date_invoice.year,
                'merged_month_year': inv.date_invoice.strftime('%b') + '-' + str(inv.date_invoice.year)
            }

            if month_year not in unique_month:
                unique_month.append(month_year)
        return unique_month

    # Get Service list
    def get_service_type_unique(self, invoice_list):
        unique_service = []
        for inv in invoice_list:
            val = {
                'service': inv.freight_booking.service_type,
                'direction': inv.freight_booking.direction
            }

            if val not in unique_service and val["service"] != False and val["direction"] != False:
                unique_service.append(val)
        unique_service.sort(key=lambda x: x["service"])
        return unique_service

    # get invoice by service type:
    def get_invoice_by_service(self, service, inv_data):
        data = []
        for inv in inv_data['data']:
            if inv.freight_booking.service_type == service['service'] and inv.freight_booking.direction == service[
                'direction']:
                data.append(inv)
        return {
            'service': service['service'],
            'direction': service['direction'],
            'data': data,
            'month': inv_data['month'],
            'year': inv_data['year'],
            'merged_month_year': inv_data['merged_month_year']
        }

    def group_invoice_vendor(self, grouped_data, invoice_list):
        grouped_invoice = []

        for inv in grouped_data['data']:
                grouped_invoice.append({
                    'date_invoice': inv.date_invoice.strftime("%d-%b-%Y"),
                    'invoice_number': inv.number,
                    'type': inv.type,
                    'booking_number': inv.freight_booking.booking_no,
                    'amount': inv.amount_total_company_signed,
                    'customer_name': inv.partner_id.name,
                    'freight_booking': inv.freight_booking
                })
        grouped_invoice = sorted(grouped_invoice, key=lambda x: (x['booking_number'], x['date_invoice']))

        processed_bookings = set()
        for gi in grouped_invoice:
            key = (gi["booking_number"], gi["type"])
            if key in processed_bookings:
                continue  # Skip if this booking number is already processed

            # Sum vendor bills for this booking number
            sum_vendor_bill = sum(
                inv_vendor.amount_total_company_signed
                for inv_vendor in invoice_list
                if gi[
                    "booking_number"] == inv_vendor.freight_booking.booking_no and inv_vendor.type == 'in_invoice' and inv_vendor.date_invoice.strftime(
                    "%b") == datetime.strptime(gi['date_invoice'], "%d-%b-%Y").strftime("%b")
            )

            sum_revenue = sum(
                inv_vendor.amount_total_company_signed
                for inv_vendor in invoice_list
                if gi[
                    "booking_number"] == inv_vendor.freight_booking.booking_no and inv_vendor.type == 'out_invoice' and inv_vendor.date_invoice.strftime(
                    "%b") == datetime.strptime(gi['date_invoice'], "%d-%b-%Y").strftime("%b")
            )

            if sum_vendor_bill > 0:
                gi["cost"] = sum_vendor_bill  # Assign cost to the first occurrence
                gi["profit"] = sum_revenue - gi["cost"]
                if sum_revenue != 0:
                    gi["margin"] = round((gi["profit"] / sum_revenue) * 100, 2)
                else:
                    gi["margin"] = 0
                processed_bookings.add(key)

        # new_grouped_invoice = []
        # processed_keys = set()
        # for inv in grouped_invoice:
        #     if inv['type'] == 'out_invoice':
        #         key = (inv['booking_number'], inv['type'])
        #         if key not in processed_keys:
        #             new_grouped_invoice.append(inv)
        #             processed_keys.add(key)
        data_in_invoice = []

        for inv in grouped_invoice:
            if inv['type']  == 'out_invoice':
                data_in_invoice.append(inv)
            elif inv['type'] == 'in_invoice':
                search_in_invoice_list = [
                    item for item in grouped_invoice
                    if item['booking_number'] == inv['booking_number'] and item['type'] == 'out_invoice' and
                       datetime.strptime(item['date_invoice'], "%d-%b-%Y").strftime("%b") ==
                       datetime.strptime(inv['date_invoice'], "%d-%b-%Y").strftime("%b")
                ]
                if len(search_in_invoice_list) == 0:
                    inv['amount'] = 0
                    inv['margin'] = 0
                    # inv["margin"] = round((inv["profit"] / inv['amount']) * 100, 2)
                    data_in_invoice.append(inv)

        data_in_invoice = sorted(data_in_invoice, key=lambda x: (x['booking_number'], x['date_invoice']))

        vals = {
            'grouped_invoice':  data_in_invoice,
            'service': grouped_data['service'],
            'direction': grouped_data['direction'],
            'month': grouped_data['month'],
            'year': grouped_data['year'],
            'merged_month_year': grouped_data['merged_month_year']
        }

        return vals

    def create_frame_dict(self,data,key):
        dict_data = []
        column = ['cost','profit','margin']
        for dt in data:
            if key not in  column:
                dict_data.append(dt[key])
            else:
                if dt.get(key) and key in  column:
                    dict_data.append(dt[key])
                else:
                    dict_data.append(0)
        return dict_data

    def sum_all_value_by_key(self,data):
        return sum(item for item in data if item != "")

    def format_numbers(self,data):
        """Formats all numeric values in a dictionary to two decimal places with a thousand separator."""
        for i, value in enumerate(data):
            if isinstance(value, (int, float)):
                data[i] = "{:,.2f}".format(value)

        return data

    @api.multi
    def get_report_invoice_wizard(self):
        get_invoice_lists = self.env["account.invoice"].search([
            ('date_invoice', '>=', self.date_from),
            ('date_invoice', '<=', self.date_to),
        ], order="date_invoice")

        # Get the month
        month_list = self.get_month_unique(get_invoice_lists)

        # Get all Service type
        service_list = self.get_service_type_unique(get_invoice_lists)

        # Group invoice by month and year
        invoice_by_month = []
        for month_dt in month_list:
            data = []
            for inv in get_invoice_lists:
                if inv.date_invoice.strftime('%b') == month_dt["month"] and inv.date_invoice.year == month_dt["year"]:
                    data.append(inv)
            invoice_by_month.append({
                'month': month_dt["month"],
                "year": month_dt['year'],
                'merged_month_year': month_dt['merged_month_year'],
                'data': data})

        # Group invoice by mon
        grouped_data = []
        for sl in service_list:
            for inv_data in invoice_by_month:
                grouped_invoice = self.get_invoice_by_service(sl, inv_data)
                grouped_data.append(grouped_invoice)

        # # Group by vendor and invoice
        cleaned_data = []
        for item in grouped_data:
            inv_by = self.group_invoice_vendor(item, get_invoice_lists)
            if len(inv_by) > 0:
                cleaned_data.append(inv_by)

        # # Cleaned Version
        cleaned_data = sorted(cleaned_data, key=lambda x: (x['year'],x['month'],x['service'],x['direction']))

        # Excel writing
        data_frame_list = []

        for month in month_list:
            data_frame_list.append(
                {'df':pd.DataFrame({}),
                 'merged_month':month['merged_month_year'],
                 'month':month['month'],
                 'year':month['year']
                 })

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            excel_path = tmp_file.name

            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for dt_frame in data_frame_list:
                    month_year = dt_frame['month'] + ' ' + str(dt_frame['year'])
                    dt_frame['df'].to_excel(writer, index=False, sheet_name=month_year, startrow=2)
                    worksheet = writer.sheets[month_year]
                    worksheet["A1"], worksheet["A1"].font = dt_frame["merged_month"], Font(bold=True)

                    cleaned_data_display = [
                        item for item in cleaned_data
                        if item['month'] == dt_frame['month'] and item['year'] == dt_frame['year']
                    ]
                    index = 0
                    revenue = 0
                    cost = 0
                    profit = 0
                    margin = 0
                    for item_clean in cleaned_data_display:
                        # Add a blank row after each direction entry
                        worksheet.cell(row=2 + index, column=1, value="")
                        index += 1

                        worksheet.cell(row=2 + index, column=1, value=f"{item_clean['service'].upper()} {item_clean['direction'].upper()}").font = Font(bold=True, size=14)
                        index += 1

                        column = 1
                        column_data = [
                            "Invoice Date","Invoice Number","Customer Name","Job Number","Revenue","Cost","Profit","Margin"
                        ]
                        for col in column_data:
                           worksheet.cell(row=2 + index, column=column, value=col).font = Font(bold=True, size=12)
                           column+=1
                        index += 1

                        # Write invoice data using data frame
                        invoice_data ={
                            "Invoice Date":self.create_frame_dict(item_clean["grouped_invoice"],"date_invoice"),
                            "Invoice Number":self.create_frame_dict(item_clean["grouped_invoice"],"invoice_number"),
                            "Customer Name":self.create_frame_dict(item_clean["grouped_invoice"],"customer_name"),
                            "Job Number":self.create_frame_dict(item_clean["grouped_invoice"],"booking_number"),
                            "Revenue":self.create_frame_dict(item_clean["grouped_invoice"],"amount"),
                            'Cost':self.create_frame_dict(item_clean["grouped_invoice"],"cost"),
                            'Profit':self.create_frame_dict(item_clean["grouped_invoice"],"profit"),
                            'Margin':self.create_frame_dict(item_clean["grouped_invoice"],"margin")
                        }

                        copy_invoice_data = copy.deepcopy(invoice_data)
                        copy_invoice_data["Revenue"] = self.format_numbers(copy_invoice_data["Revenue"])
                        copy_invoice_data["Cost"] = self.format_numbers(copy_invoice_data["Cost"])
                        copy_invoice_data["Profit"] = self.format_numbers(copy_invoice_data["Profit"])
                        copy_invoice_data["Margin"] = self.format_numbers(copy_invoice_data["Margin"])


                        df = pd.DataFrame(copy_invoice_data)
                        revenue += self.sum_all_value_by_key(invoice_data["Revenue"])
                        cost += self.sum_all_value_by_key(invoice_data["Cost"])
                        profit += self.sum_all_value_by_key(invoice_data["Profit"])
                        margin += self.sum_all_value_by_key(invoice_data["Margin"])

                        for row_idx, row in df.iterrows():
                            for col_idx, col in enumerate(column_data, start=1):
                               cell= worksheet.cell(row=2 + index, column=col_idx, value=row[col])  # Write invoice data
                               col_letter = cell.column_letter  # Get the column letter
                               if col in ["Invoice Date","Revenue","Cost","Profit","Margin"]:
                                   cell.alignment = Alignment(horizontal="right")
                               if col in ["Invoice Number","Job Number"]:
                                   worksheet.column_dimensions[col_letter].width = 25
                               if col in ["Invoice Date"]:
                                   worksheet.column_dimensions[col_letter].width = 18
                               if col in ["Customer Name",]:
                                   worksheet.column_dimensions[col_letter].width = 45
                               if col in ["Revenue","Cost","Profit","Margin"]:
                                   worksheet.column_dimensions[col_letter].width = 20

                               cell.font = Font(size=12)
                            index += 1  # Move to the next row

                        index += 1

                    # Display sum of total Revenue until margin
                    total_column = [f"MONTH {dt_frame['month'].upper()}", "", "", "", "Total Revenue", "Total Cost", "Profit",
                                    "Margin"]

                    sum_results = ["", "", "", "", "{:,.2f}".format(revenue), "{:,.2f}".format(cost), "{:,.2f}".format(profit), "{:,.2f}".format(margin)]

                    for row_offset, data in enumerate([total_column, sum_results]):
                        for col_idx, col in enumerate(data, start=1):
                           cell= worksheet.cell(row=2 + index + row_offset, column=col_idx, value=col)

                           # Align numeric values in the sum_results row to the right
                           if isinstance(col, (int, float)):
                               cell.alignment = Alignment(horizontal="right")
                               cell.font = Font(size=12)
                           elif col != '' and isinstance(col,(str)):
                               cell.font = Font(bold=True,size=12)
                    index += 2


        with open(excel_path, 'rb') as file:
            file_data = base64.b64encode(file.read())

        # Attach file data to the wizard
        self.file_data = file_data
        current_date = datetime.now().strftime("%d-%m-%y")
        self.file_name = f"{current_date}_invoiced_report.xlsx"

        # Open a download wizard view
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }
